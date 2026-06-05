# CONFIGURATION — edit these two lines before running
INPUT_FOLDER = r"USB0\2025-11"  # path to the month folder containing TXT files
OUTPUT_FILE  = r"November_generated.xlsx"  # output workbook path

import re
from datetime import datetime, date, time
from pathlib import Path
import openpyxl
from openpyxl.utils import get_column_letter


_WALLCLOCK_RE = re.compile(r'^(\d{4}-\d{2}-\d{2}) (\d{2}:\d{2}:\d{2})(.*)', re.DOTALL)

# parses one raw line from a sensor log
def parse_line(line):
    m = _WALLCLOCK_RE.match(line)
    if not m:
        return None

    wc_date = datetime.strptime(m.group(1), '%Y-%m-%d').date()
    wc_time = datetime.strptime(m.group(2), '%H:%M:%S').time()
    remainder = m.group(3).strip()

    if not remainder:
        return (wc_date, wc_time, 'blank', None)

    if remainder.startswith('Date'):
        return (wc_date, wc_time, 'header', None)

    # data line: DATA_DATE DATA_TIME RECORD# T01 T02 T03 T04 T05 T06 T07 T08
    parts = remainder.split()
    if len(parts) < 9:
        return (wc_date, wc_time, 'blank', None)

    fields = []
    for raw in parts[3:9]:  # indices 3–8 = T01–T06
        if raw.lower() == 'nan':
            fields.append(None)
        else:
            try:
                fields.append(float(raw))
            except ValueError:
                fields.append(None)

    return (wc_date, wc_time, 'data', fields)

# reads all of the logs for the month (alphabetically), returns the first header's timestamp + list of rows in order
def load_month_folder(folder_path):
    folder = Path(folder_path)
    txt_files = sorted(folder.glob('RDL_*_USB0.txt'))

    header_wc = None
    rows = []

    for txt_file in txt_files:
        with open(txt_file, 'r', encoding='utf-8', errors='replace') as fh:
            for line in fh:
                result = parse_line(line.rstrip('\r\n'))
                if result is None:
                    continue
                wc_date, wc_time, kind, fields = result
                if kind == 'header' and header_wc is None:
                    header_wc = (wc_date, wc_time)
                elif kind == 'data':
                    rows.append((wc_date, wc_time) + tuple(fields))

    return header_wc, rows


_COL_LABELS = [
    'T01', 'T02', 'Under Modules inside Pots',
    'T03', 'T04', 'Under Modules',
    'T05', 'T06', 'Control',
]

# builds and saves the formatted workbook
def write_excel(header_wc, rows, output_path, sheet_name):
    if not rows:
        raise ValueError("rows must not be empty")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name
    ws.sheet_view.zoomScale = 70

    ws.column_dimensions['A'].width = 12.88671875
    ws.column_dimensions['E'].width = 23.44140625
    ws.column_dimensions['H'].width = 13.5546875
    ws.column_dimensions['K'].width = 7.44140625

    # Row 1: header
    hdr_date, hdr_time = header_wc
    ws['A1'] = datetime.combine(hdr_date, time(0, 0, 0))
    ws['A1'].number_format = 'mm-dd-yy'
    ws['B1'] = hdr_time
    ws['B1'].number_format = 'h:mm:ss'
    for col_idx, label in enumerate(_COL_LABELS, start=3):
        ws.cell(row=1, column=col_idx, value=label)

    # rows 2-N: data
    for i, row_data in enumerate(rows, start=2):
        r_date, r_time, t01, t02, t03, t04, t05, t06 = row_data
        a = ws.cell(i, 1, datetime.combine(r_date, time(0, 0, 0)))
        a.number_format = 'mm-dd-yy'
        b = ws.cell(i, 2, r_time)
        b.number_format = 'h:mm:ss'
        ws.cell(i, 3, t01)
        ws.cell(i, 4, t02)
        ws.cell(i, 5, f'=AVERAGE(C{i}:D{i})')
        ws.cell(i, 6, t03)
        ws.cell(i, 7, t04)
        ws.cell(i, 8, f'=AVERAGE(F{i}:G{i})')
        ws.cell(i, 9, t05)
        ws.cell(i, 10, t06)
        ws.cell(i, 11, f'=AVERAGE(I{i}:J{i})')

    # summary row
    last_data_row = len(rows) + 1   # header is row 1, first data is row 2
    summary_row   = last_data_row + 1
    for col_idx in range(3, 12):    # C-K
        col_letter = get_column_letter(col_idx)
        ws.cell(summary_row, col_idx,
                f'=AVERAGE({col_letter}2:{col_letter}{last_data_row})')

    wb.save(output_path)


def main():
    folder = Path(INPUT_FOLDER)
    if not folder.is_dir():
        raise SystemExit(f"Error: INPUT_FOLDER not found: {folder.resolve()}")

    print(f"Reading TXT files from: {folder.resolve()}")
    header_wc, rows = load_month_folder(folder)

    if not rows:
        raise SystemExit("Error: no data rows found — check INPUT_FOLDER and file naming.")
    if header_wc is None:
        raise SystemExit("Error: no header line found in any TXT file.")

    first_date = rows[0][0]
    sheet_name  = first_date.strftime("%B")   # e.g. "November"

    print(f"Loaded {len(rows)} data rows  ({first_date} -> {rows[-1][0]})")
    print(f"Writing: {Path(OUTPUT_FILE).resolve()}")
    write_excel(header_wc, rows, OUTPUT_FILE, sheet_name)
    print("Done.")


if __name__ == "__main__":
    main()
